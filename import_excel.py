import pandas as pd
import os
from typing import Optional, Dict, Any, List
from werkzeug.utils import secure_filename
import uuid
from datetime import datetime
import numpy as np

# Cấu hình upload
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}
MAX_FILE_SIZE = 100 * 1024 * 1024  # 100MB

def _safe_dict_records(df: pd.DataFrame) -> List[Dict[str, Any]]:
    """
    Safely convert DataFrame to dict records, handling non-JSON serializable types
    """
    try:
        # Replace all NaN/None values with empty string
        df_clean = df.copy()
        
        # Convert datetime columns to string
        for col in df_clean.columns:
            if pd.api.types.is_datetime64_any_dtype(df_clean[col]):
                df_clean[col] = df_clean[col].dt.strftime('%Y-%m-%d %H:%M:%S')
            elif pd.api.types.is_numeric_dtype(df_clean[col]):
                # Convert numeric NaN to 0 or empty string
                df_clean[col] = df_clean[col].fillna(0)
            else:
                # Convert all other types to string and handle NaN
                df_clean[col] = df_clean[col].astype(str).replace('nan', '')
        
        # Fill any remaining NaN values
        df_clean = df_clean.fillna('')
        
        return df_clean.to_dict('records')
    except Exception as e:
        print(f"Error in _safe_dict_records: {e}")
        # Fallback: convert everything to string
        try:
            df_str = df.fillna('').astype(str)
            return df_str.to_dict('records')
        except:
            return []

def ensure_upload_folder():
    """Tạo thư mục uploads nếu chưa tồn tại"""
    if not os.path.exists(UPLOAD_FOLDER):
        os.makedirs(UPLOAD_FOLDER)
        print(f"Đã tạo thư mục: {UPLOAD_FOLDER}")

def allowed_file(filename: str) -> bool:
    """Kiểm tra file có phần mở rộng hợp lệ không"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def generate_unique_filename(filename: str) -> str:
    """Tạo tên file duy nhất để tránh trùng lặp"""
    secure_name = secure_filename(filename)
    name, ext = os.path.splitext(secure_name)
    unique_id = str(uuid.uuid4())[:8]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{name}_{timestamp}_{unique_id}{ext}"

def save_uploaded_file(file_storage, custom_filename: Optional[str] = None) -> Dict[str, Any]:
    """
    Lưu file được upload và trả về thông tin file
    
    Args:
        file_storage: File object từ Flask request
        custom_filename (str, optional): Tên file tùy chỉnh
    
    Returns:
        Dict[str, Any]: Thông tin về file đã lưu
    """
    try:
        ensure_upload_folder()
        
        if file_storage.filename == '':
            raise ValueError("Không có file nào được chọn")
        
        if not allowed_file(file_storage.filename):
            raise ValueError(f"Loại file không được hỗ trợ. Chỉ chấp nhận: {', '.join(ALLOWED_EXTENSIONS)}")
        
        # Tạo tên file an toàn và duy nhất
        if custom_filename:
            filename = secure_filename(custom_filename)
            if not allowed_file(filename):
                filename = generate_unique_filename(file_storage.filename)
        else:
            filename = generate_unique_filename(file_storage.filename)
        
        file_path = os.path.join(UPLOAD_FOLDER, filename)
        
        # Lưu file
        file_storage.save(file_path)
        
        # Lấy thông tin file
        file_size = os.path.getsize(file_path)
        
        result = {
            'success': True,
            'filename': filename,
            'file_path': file_path,
            'file_size': file_size,
            'file_size_mb': round(file_size / (1024 * 1024), 2),
            'upload_time': datetime.now().isoformat(),
            'original_filename': file_storage.filename
        }
        
        print(f"Đã lưu file thành công: {file_path} ({result['file_size_mb']} MB)")
        return result
        
    except Exception as e:
        print(f"Lỗi khi lưu file: {str(e)}")
        return {
            'success': False,
            'error': str(e)
        }

def process_uploaded_excel(file_info: Dict[str, Any]) -> Dict[str, Any]:
    """
    Xử lý và phân tích file Excel đã upload
    
    Args:
        file_info (Dict[str, Any]): Thông tin file từ save_uploaded_file()
    
    Returns:
        Dict[str, Any]: Kết quả phân tích file Excel
    """
    try:
        if not file_info.get('success'):
            raise ValueError(file_info.get('error', 'File upload không thành công'))
        
        file_path = file_info['file_path']
        
        # Lấy thông tin Excel
        excel_info = get_excel_info(file_path)
        
        # Đọc sheet đầu tiên
        df = read_excel_file(file_path)
        
        # Tạo summary
        result = {
            'success': True,
            'file_info': file_info,
            'excel_info': excel_info,
            'data_summary': {
                'total_rows': len(df),
                'total_columns': len(df.columns),
                'columns': list(df.columns),
                'dtypes': {col: str(dtype) for col, dtype in df.dtypes.items()},
                'null_counts': {col: int(count) for col, count in df.isnull().sum().items()},
                'sample_data': _safe_dict_records(df.head(5)) if not df.empty else []
            }
        }
        
        print(f"Đã phân tích file Excel thành công")
        print(f"Dữ liệu: {result['data_summary']['total_rows']} dòng, {result['data_summary']['total_columns']} cột")
        
        return result
        
    except Exception as e:
        print(f"Lỗi khi phân tích Excel: {str(e)}")
        return {
            'success': False,
            'error': str(e),
            'file_info': file_info
        }

def get_uploaded_files() -> List[Dict[str, Any]]:
    """Lấy danh sách tất cả files đã upload"""
    try:
        ensure_upload_folder()
        files = []
        
        for filename in os.listdir(UPLOAD_FOLDER):
            if allowed_file(filename):
                file_path = os.path.join(UPLOAD_FOLDER, filename)
                file_stat = os.stat(file_path)
                
                files.append({
                    'filename': filename,
                    'file_path': file_path,
                    'file_size_mb': round(file_stat.st_size / (1024 * 1024), 2),
                    'upload_time': datetime.fromtimestamp(file_stat.st_mtime).isoformat(),
                    'is_excel': filename.lower().endswith(('.xlsx', '.xls'))
                })
        
        # Sắp xếp theo thời gian upload (mới nhất trước)
        files.sort(key=lambda x: x['upload_time'], reverse=True)
        return files
        
    except Exception as e:
        print(f"Lỗi khi lấy danh sách file: {str(e)}")
        return []

def delete_uploaded_file(filename: str) -> bool:
    """Xóa file đã upload"""
    try:
        file_path = os.path.join(UPLOAD_FOLDER, secure_filename(filename))
        if os.path.exists(file_path):
            os.remove(file_path)
            print(f"Đã xóa file: {filename}")
            return True
        else:
            print(f"File không tồn tại: {filename}")
            return False
    except Exception as e:
        print(f"Lỗi khi xóa file: {str(e)}")
        return False

def read_excel_file(file_path: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
    """
    Đọc file Excel và trả về DataFrame
    
    Args:
        file_path (str): Đường dẫn đến file Excel
        sheet_name (str, optional): Tên sheet cần đọc. Nếu None sẽ đọc sheet đầu tiên
    
    Returns:
        pd.DataFrame: Dữ liệu từ file Excel
    
    Raises:
        FileNotFoundError: Nếu file không tồn tại
        Exception: Nếu có lỗi khi đọc file
    """
    try:
        # Kiểm tra file có tồn tại không
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File không tồn tại: {file_path}")
        
        # Đọc file Excel
        if sheet_name:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            print(f"Đã đọc thành công sheet '{sheet_name}' từ file: {file_path}")
        else:
            df = pd.read_excel(file_path)
            print(f"Đã đọc thành công file: {file_path}")
        
        return df
    
    except Exception as e:
        print(f"Lỗi khi đọc file Excel: {str(e)}")
        raise

def get_excel_info(file_path: str) -> Dict[str, Any]:
    """
    Lấy thông tin về file Excel (số sheet, tên các sheet, etc.)
    
    Args:
        file_path (str): Đường dẫn đến file Excel
    
    Returns:
        Dict[str, Any]: Thông tin về file Excel
    """
    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File không tồn tại: {file_path}")
        
        # Đọc tất cả sheet names
        excel_file = pd.ExcelFile(file_path)
        
        info = {
            'file_path': file_path,
            'sheet_names': excel_file.sheet_names,
            'number_of_sheets': len(excel_file.sheet_names),
            'file_size_mb': round(os.path.getsize(file_path) / (1024 * 1024), 2)
        }
        
        return info
    
    except Exception as e:
        print(f"Lỗi khi lấy thông tin file Excel: {str(e)}")
        raise

def display_dataframe_info(df: pd.DataFrame, sample_rows: int = 5) -> None:
    """
    Hiển thị thông tin về DataFrame
    
    Args:
        df (pd.DataFrame): DataFrame cần hiển thị thông tin
        sample_rows (int): Số dòng mẫu cần hiển thị
    """
    print("=" * 50)
    print("THÔNG TIN DATAFRAME:")
    print("=" * 50)
    print(f"Kích thước: {df.shape[0]} dòng, {df.shape[1]} cột")
    print(f"Tên các cột: {list(df.columns)}")
    print(f"Kiểu dữ liệu các cột:")
    print(df.dtypes)
    print(f"\n{sample_rows} dòng đầu tiên:")
    print(df.head(sample_rows))
    
    # Kiểm tra dữ liệu null
    null_counts = df.isnull().sum()
    if null_counts.any():
        print(f"\nSố lượng giá trị null trong mỗi cột:")
        print(null_counts[null_counts > 0])
    else:
        print("\nKhông có giá trị null nào trong dữ liệu.")

def read_specific_columns(file_path: str, columns: list, sheet_name: Optional[str] = None) -> pd.DataFrame:
    """
    Đọc chỉ những cột cụ thể từ file Excel
    
    Args:
        file_path (str): Đường dẫn đến file Excel
        columns (list): Danh sách tên cột cần đọc
        sheet_name (str, optional): Tên sheet cần đọc
    
    Returns:
        pd.DataFrame: DataFrame chỉ chứa các cột được chọn
    """
    try:
        df = read_excel_file(file_path, sheet_name)
        
        # Kiểm tra xem các cột có tồn tại không
        missing_cols = [col for col in columns if col not in df.columns]
        if missing_cols:
            print(f"Cảnh báo: Các cột không tồn tại: {missing_cols}")
            columns = [col for col in columns if col in df.columns]
        
        if not columns:
            raise ValueError("Không có cột nào hợp lệ để đọc")
        
        return df[columns]
    
    except Exception as e:
        print(f"Lỗi khi đọc các cột cụ thể: {str(e)}")
        raise

# Demo và test functions
if __name__ == "__main__":
    print("=" * 60)
    print("EXCEL UPLOAD & DOWNLOAD SYSTEM")
    print("=" * 60)
    
    # Kiểm tra files có sẵn trong uploads
    print("Kiểm tra files trong thư mục uploads...")
    uploaded_files = get_uploaded_files()
    
    if uploaded_files:
        print(f"Tìm thấy {len(uploaded_files)} file(s):")
        for i, file in enumerate(uploaded_files, 1):
            print(f"  {i}. {file['filename']} ({file['file_size_mb']} MB)")
        
        # Test với file đầu tiên
        test_file = uploaded_files[0]
        print(f"\nTesting với file: {test_file['filename']}")
        
        try:
            # Test đọc info
            info = get_excel_info(test_file['file_path'])
            print(f"Thông tin Excel:")
            print(f"  - Số sheets: {info['number_of_sheets']}")
            print(f"  - Tên sheets: {info['sheet_names']}")
            print(f"  - Kích thước: {info['file_size_mb']} MB")
            
            # Test đọc dữ liệu
            df = read_excel_file(test_file['file_path'])
            print(f"\nDữ liệu:")
            print(f"  - Số dòng: {len(df)}")
            print(f"  - Số cột: {len(df.columns)}")
            print(f"  - Các cột: {list(df.columns)}")
            
            if not df.empty:
                print(f"\nDữ liệu mẫu (3 dòng đầu):")
                print(df.head(3).to_string(index=False))
            
            print(f"\nTất cả functions hoạt động bình thường!")
            
        except Exception as e:
            print(f"Lỗi khi test: {str(e)}")
    
    else:
        print("Thư mục uploads trống.")
        print("\nHƯỚNG DẪN SỬ DỤNG:")
        print("1. Chạy ứng dụng web: python app.py")
        print("2. Truy cập: http://localhost:5000/excel-upload")
        print("3. Upload file Excel của bạn")
        print("4. Download và tự động xóa file")
    
    print("\n" + "=" * 60)

def get_images_from_drive_folder(drive_handler, folder_name, max_images=None):
    """
    Lấy ảnh đầu tiên từ Google Drive folder
    
    Args:
        drive_handler: GoogleDriveUploader instance
        folder_name: Tên folder cần tìm
        max_images: Không sử dụng (chỉ lấy 1 ảnh đầu tiên)
    
    Returns:
        dict: Thông tin ảnh đầu tiên hoặc None nếu không tìm thấy
    """
    try:
        # Tìm folder theo tên
        folder_id = drive_handler.find_folder(folder_name)
        
        if not folder_id:
            print(f"Folder '{folder_name}' not found on Google Drive")
            return []
        
        print(f"Found folder '{folder_name}' with ID: {folder_id}")
        
        # List files trong folder
        query = f"'{folder_id}' in parents and mimeType contains 'image/'"
        
        results = drive_handler.service.files().list(
            q=query,
            orderBy='name',
            fields="files(id, name, webViewLink, webContentLink)"
        ).execute()
        
        files = results.get('files', [])
        
        if not files:
            print(f"No images found in folder '{folder_name}'")
            return None
        
        # Chỉ lấy ảnh đầu tiên
        first_file = files[0]
        print(f"Found {len(files)} images in folder '{folder_name}', using first image: {first_file['name']}")
        
        # Return thông tin ảnh đầu tiên (sử dụng view link)
        file_id = first_file['id']
        view_url = f"https://drive.google.com/file/d/{file_id}/view"
        
        return {
            'name': first_file['name'],
            'url': view_url,
            'id': file_id
        }
        
    except Exception as e:
        print(f"Error getting images from folder '{folder_name}': {str(e)}")
        return None

def fill_excel_with_data(filename, products, column='E', start_row=7, fill_mode='repeat', sku_column='A', batch_column='CU', image_column='T', fill_images_from_drive=False):
    """
    Fill dữ liệu từ database vào file Excel
    
    Args:
        filename (str): Tên file Excel cần fill
        products (list): Danh sách sản phẩm từ database
        column (str): Cột cần fill (mặc định 'E')
        start_row (int): Dòng bắt đầu fill (mặc định 7)
    
    Returns:
        dict: Kết quả fill dữ liệu
    """
    try:
        import openpyxl
        from datetime import datetime
        
        ensure_upload_folder()
        file_path = os.path.join(UPLOAD_FOLDER, filename)
        
        if not os.path.exists(file_path):
            return {
                'success': False,
                'error': f'File không tồn tại: {filename}'
            }
        
        # Đọc file Excel
        print(f"Đang đọc file: {file_path}")
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active
        
        print(f"Sheet name: {worksheet.title}")
        print(f"Sheet dimensions: {worksheet.max_row} rows x {worksheet.max_column} columns")
        
        # Kiểm tra nội dung cột E trước khi fill
        print(f"Checking existing data in column {column} from row {start_row-2} to {start_row+2}:")
        for check_row in range(max(1, start_row-2), start_row+3):
            cell_val = worksheet[f"{column}{check_row}"].value
            print(f"  {column}{check_row}: {cell_val}")
        
        # Nếu không có sản phẩm thì không làm gì
        if not products:
            print("Không có sản phẩm để fill!")
            return {
                'success': False,
                'error': 'Không có sản phẩm trong collection'
            }
        
        # Lọc sản phẩm có tên
        valid_products = [product for product in products if product.get('name', '').strip()]
        
        if not valid_products:
            print("Không có sản phẩm nào có tên!")
            return {
                'success': False,
                'error': 'Không có sản phẩm nào có tên trong collection'
            }
        
        print(f"Found {len(valid_products)} valid products to fill")
        print(f"Fill mode: {fill_mode}")
        print(f"Product name column: {column}, SKU column: {sku_column}, Batch ID column: {batch_column}")
        
        # Generate unique Batch ID cho toàn bộ file này
        from walmart import generate_batch_id
        batch_id = generate_batch_id()
        print(f"Generated Batch ID for this file: {batch_id}")
        
        # Initialize Google Drive handler if needed for images
        drive_handler = None
        if fill_images_from_drive:
            try:
                from google_drive_handler import GoogleDriveUploader
                drive_handler = GoogleDriveUploader()
                print(f"Google Drive connection established for image retrieval")
            except Exception as e:
                print(f"Warning: Could not connect to Google Drive for images: {str(e)}")
                fill_images_from_drive = False
        
        # Tự động phát hiện dòng cuối thực sự có dữ liệu
        last_data_row = worksheet.max_row
        print(f"Excel max_row: {last_data_row}")
        
        # Scan từ cuối lên để tìm dòng cuối thực sự có dữ liệu (không phải dòng trống)
        actual_last_row = last_data_row
        for row_num in range(last_data_row, 0, -1):
            has_data = False
            for col_idx in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=row_num, column=col_idx).value
                if cell_value is not None and str(cell_value).strip() != '':
                    has_data = True
                    break
            if has_data:
                actual_last_row = row_num
                break
        
        print(f"Actual last row with data: {actual_last_row}")
        
        # Tính số dòng cần fill từ start_row đến actual_last_row
        if actual_last_row < start_row:
            # Nếu file quá ngắn, fill ít nhất 10 dòng
            fill_until_row = start_row + 9  # fill từ start_row đến start_row+9 (10 dòng)
            print(f"File ngắn, sẽ fill 10 dòng từ {start_row} đến {fill_until_row}")
        else:
            fill_until_row = actual_last_row
            print(f"Sẽ fill từ dòng {start_row} đến dòng {fill_until_row}")
        
        original_rows_to_fill = fill_until_row - start_row + 1
        print(f"Original rows to fill: {original_rows_to_fill}")
        
        # Prepare image mapping if needed
        product_images = {}
        if fill_images_from_drive and drive_handler:
            print(f"Preparing image mapping from Google Drive...")
            
            # Get unique product names for image mapping
            unique_product_names = set()
            for product in valid_products:
                name = product.get('name', '').strip()
                if name:
                    unique_product_names.add(name)
            
            print(f"Found {len(unique_product_names)} unique products for image mapping")
            
            # Get images for each unique product
            for product_name in unique_product_names:
                # Convert product name to folder name (same logic as Google Drive upload)
                import re
                clean_name = re.sub(r'[^\w\s-]', '', product_name)
                folder_name = re.sub(r'[-\s]+', '_', clean_name).strip('_')
                
                # Get first image from Google Drive folder
                first_image = get_images_from_drive_folder(drive_handler, folder_name)
                if first_image:
                    product_images[product_name] = first_image
                    print(f"📁 '{folder_name}': 1 image → {first_image['name']}")
            
            print(f"Image mapping completed: {len(product_images)} products have images")
        
        filled_count = 0
        
        if fill_mode == 'duplicate':
            # DUPLICATE MODE: Mỗi sản phẩm unique sẽ có 1 nhóm dòng riêng
            unique_products = []
            seen_names = set()
            for product in valid_products:
                name = product.get('name', '').strip()
                if name and name not in seen_names:
                    unique_products.append(product)
                    seen_names.add(name)
            
            print(f"Found {len(unique_products)} unique products")
            
            current_row = start_row
            
            for group_index, unique_product in enumerate(unique_products):
                product_name = unique_product.get('name', '').strip()
                print(f"Processing group {group_index + 1}: '{product_name}' (rows {current_row}-{current_row + original_rows_to_fill - 1})")
                
                # Copy structure từ original rows nếu không phải group đầu tiên
                if group_index > 0:
                    # Copy tất cả cells từ original group (start_row -> fill_until_row) 
                    # sang group hiện tại (current_row -> current_row + original_rows_to_fill - 1)
                    for row_offset in range(original_rows_to_fill):
                        source_row = start_row + row_offset
                        target_row = current_row + row_offset
                        
                        # Copy tất cả cells trong dòng (trừ cột product name, SKU, và batch ID)
                        for col_idx in range(1, worksheet.max_column + 1):
                            col_letter = worksheet.cell(row=1, column=col_idx).column_letter
                            if col_letter not in [column, sku_column, batch_column]:  # Không copy các cột sẽ fill mới
                                source_cell = worksheet.cell(row=source_row, column=col_idx)
                                target_cell = worksheet.cell(row=target_row, column=col_idx)
                                
                                # Copy value, style, format
                                target_cell.value = source_cell.value
                                if hasattr(source_cell, '_style'):
                                    target_cell._style = source_cell._style
                
                # Fill product name, SKU, và Batch ID vào tất cả dòng của group này
                for row_offset in range(original_rows_to_fill):
                    target_row = current_row + row_offset
                    
                    # Fill product name
                    product_cell_address = f"{column}{target_row}"
                    worksheet[product_cell_address] = product_name
                    
                    # Generate và fill unique SKU  
                    from walmart import generate_unique_sku, mark_sku_used
                    unique_sku = generate_unique_sku()
                    sku_cell_address = f"{sku_column}{target_row}"
                    worksheet[sku_cell_address] = unique_sku
                    mark_sku_used(unique_sku)
                    
                    # Fill Batch ID (cùng giá trị cho tất cả dòng)
                    batch_cell_address = f"{batch_column}{target_row}"
                    worksheet[batch_cell_address] = batch_id
                    
                    # Fill image from Google Drive if available (same image for all rows)
                    if fill_images_from_drive and product_name in product_images:
                        image_data = product_images[product_name]
                        image_url = image_data['url']
                        image_cell_address = f"{image_column}{target_row}"
                        worksheet[image_cell_address] = image_url
                        
                        if filled_count <= 10 or filled_count % 50 == 0:
                            print(f"  Image: {image_cell_address} = {image_data['name']}")
                    
                    filled_count += 1
                    
                    if filled_count <= 10 or filled_count % 50 == 0:
                        print(f"Fill #{filled_count}: {product_cell_address} = '{product_name}', {sku_cell_address} = '{unique_sku}', {batch_cell_address} = '{batch_id}' (group {group_index + 1})")
                
                # Move to next group
                current_row += original_rows_to_fill
            
            total_filled_rows = len(unique_products) * original_rows_to_fill
            print(f"Duplicate mode completed: {len(unique_products)} groups × {original_rows_to_fill} rows = {total_filled_rows} total rows")
            
        else:
            # REPEAT MODE: Lặp lại danh sách sản phẩm như cũ
            current_row = start_row
            product_index = 0
            
            print(f"Starting repeat mode: fill {original_rows_to_fill} rows from row {start_row} to {fill_until_row}")
            
            while filled_count < original_rows_to_fill and current_row <= fill_until_row:
                # Lấy sản phẩm theo vòng lặp (cycle through products)
                product = valid_products[product_index % len(valid_products)]
                product_name = product.get('name', '').strip()
                
                if product_name:
                    # Fill product name
                    product_cell_address = f"{column}{current_row}"
                    worksheet[product_cell_address] = product_name
                    
                    # Generate và fill unique SKU
                    from walmart import generate_unique_sku, mark_sku_used
                    unique_sku = generate_unique_sku()
                    sku_cell_address = f"{sku_column}{current_row}"
                    worksheet[sku_cell_address] = unique_sku
                    mark_sku_used(unique_sku)
                    
                    # Fill Batch ID (cùng giá trị cho tất cả dòng)
                    batch_cell_address = f"{batch_column}{current_row}"
                    worksheet[batch_cell_address] = batch_id
                    
                    # Fill image from Google Drive if available (same image for all rows)
                    if fill_images_from_drive and product_name in product_images:
                        image_data = product_images[product_name]
                        image_url = image_data['url']
                        image_cell_address = f"{image_column}{current_row}"
                        worksheet[image_cell_address] = image_url
                        
                        if filled_count <= 5 or filled_count % 50 == 0:
                            print(f"  Image: {image_cell_address} = {image_data['name']}")
                    
                    filled_count += 1
                    
                    # Verify data was written
                    verify_product_value = worksheet[product_cell_address].value
                    verify_sku_value = worksheet[sku_cell_address].value
                    verify_batch_value = worksheet[batch_cell_address].value
                    
                    # Debug output for first few and every 50th item
                    if filled_count <= 5 or filled_count % 50 == 0:
                        print(f"Fill #{filled_count}: {product_cell_address} = '{product_name}' (verified: '{verify_product_value}'), {sku_cell_address} = '{unique_sku}' (verified: '{verify_sku_value}'), {batch_cell_address} = '{batch_id}' (verified: '{verify_batch_value}') [product #{product_index % len(valid_products) + 1}]")
                    
                    current_row += 1
                    product_index += 1
                else:
                    # Skip sản phẩm không có tên
                    product_index += 1
                    continue
        
        # Kiểm tra lại sau khi fill
        print(f"After filling, checking column {column} from row {start_row} to {start_row+min(5, filled_count)}:")
        for check_row in range(start_row, start_row + min(5, filled_count) + 1):
            cell_val = worksheet[f"{column}{check_row}"].value
            print(f"  {column}{check_row}: {cell_val}")
        
        # Tạo tên file mới
        name_part, ext = os.path.splitext(filename)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        new_filename = f"{name_part}_filled_{timestamp}{ext}"
        new_file_path = os.path.join(UPLOAD_FOLDER, new_filename)
        
        print(f"Saving to: {new_file_path}")
        
        # Lưu file mới
        workbook.save(new_file_path)
        workbook.close()
        
        # Verify file was saved and check size
        if os.path.exists(new_file_path):
            file_size = os.path.getsize(new_file_path)
            print(f"File saved successfully: {new_file_path} ({file_size} bytes)")
            
            # Quick verification by reopening file
            verify_workbook = openpyxl.load_workbook(new_file_path)
            verify_worksheet = verify_workbook.active
            sample_cell = verify_worksheet[f"{column}{start_row}"]
            print(f"Verification - {column}{start_row} in saved file: '{sample_cell.value}'")
            verify_workbook.close()
        else:
            print(f"ERROR: File was not saved!")
        
        print(f"Đã lưu file mới: {new_file_path}")
        print(f"Đã fill {filled_count} sản phẩm")
        
        # Mark Batch ID as used
        from walmart import mark_batch_used
        mark_batch_used(batch_id, filled_count)
        
        if fill_mode == 'duplicate':
            unique_count = len(set(p.get('name', '').strip() for p in valid_products if p.get('name', '').strip()))
            base_message = f'Chế độ duplicate: {unique_count} sản phẩm unique × {original_rows_to_fill} dòng = {filled_count} tổng dòng đã fill (Product Name + SKU + Batch ID: {batch_id}'
            if fill_images_from_drive:
                image_count = len(product_images)
                base_message += f' + Images từ {image_count} Google Drive folders'
            message = base_message + ')'
            end_row = start_row + filled_count - 1
        else:
            base_message = f'Chế độ repeat: Đã fill {filled_count} dòng vào cột {column} (Product Name), {sku_column} (SKU) và {batch_column} (Batch ID: {batch_id})'
            if fill_images_from_drive:
                image_count = len(product_images)
                base_message += f', {image_column} (Images từ {image_count} Google Drive folders)'
            message = base_message + f' từ dòng {start_row} đến {fill_until_row}'
            end_row = fill_until_row
        
        return {
            'success': True,
            'filled_count': filled_count,
            'new_filename': new_filename,
            'column': column,
            'sku_column': sku_column,
            'batch_column': batch_column,
            'batch_id': batch_id,
            'image_column': image_column if fill_images_from_drive else None,
            'images_filled': len(product_images) if fill_images_from_drive else 0,
            'start_row': start_row,
            'end_row': end_row,
            'fill_mode': fill_mode,
            'total_rows_detected': original_rows_to_fill,
            'message': message
        }
        
    except Exception as e:
        print(f"Error in fill_excel_with_data: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            'success': False,
            'error': f'Lỗi khi fill dữ liệu: {str(e)}'
        }
